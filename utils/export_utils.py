import openpyxl
from openpyxl.styles import Font
from fpdf import FPDF
import math


def exportar_productos_a_excel(productos, ruta_archivo):
    """
    Exporta la lista de productos con variantes a un archivo Excel (.xlsx).
    Cada variante ocupa una fila.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos y Variantes"

    headers = [
        "ID Producto", "Nombre", "Descripción", "Precio Base",
        "Proveedor", "Categoría", "Talla", "Color", "Stock", "SKU"
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for prod in productos:
        for var in getattr(prod, "variantes", []):
            fila = [
                prod.id_producto,
                prod.nombre,
                prod.descripcion,
                prod.precio_base,
                getattr(prod, "proveedor_nombre", ""),
                getattr(prod, "categoria_nombre", ""),
                var.talla,
                var.color,
                var.stock,
                var.sku
            ]
            ws.append(fila)

    for column_cells in ws.columns:
        max_len = max(len(str(c.value)) for c in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2

    wb.save(ruta_archivo)


def exportar_productos_a_pdf(productos, ruta_archivo):
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()

    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Productos con Variantes", border=0, ln=1, align="C")
    pdf.ln(4)

    # Definir encabezados y anchos de columna
    headers = [
        "ID", "Nombre", "Descripción", "Precio",
        "Proveedor", "Categoría", "Talla", "Color", "Stock", "SKU"
    ]
    col_widths = [12, 40, 55, 18, 35, 35, 18, 18, 15, 28]

    line_height = 6 

    pdf.set_font("Arial", "B", 9)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, h, border=1, align='C') 
    pdf.ln()

    pdf.set_font("Arial", size=8)
    
    for prod in productos:
        for var in getattr(prod, "variantes", []):
            row_data = [
                str(prod.id_producto),
                prod.nombre,
                prod.descripcion or "",
                f"${prod.precio_base:.2f}",
                getattr(prod, "proveedor_nombre", ""),
                getattr(prod, "categoria_nombre", ""),
                var.talla,
                var.color,
                str(var.stock),
                var.sku or ""
            ]

            max_row_height = line_height 

            for i, txt in enumerate(row_data):
                text_content = str(txt)
                col_width = col_widths[i]
                
                text_width = pdf.get_string_width(text_content)

                if text_width > col_width:
                    num_lines = math.ceil(text_width / col_width)
                    cell_required_height = num_lines * line_height
                    max_row_height = max(max_row_height, cell_required_height)
                else:
                    max_row_height = max(max_row_height, line_height)

            max_row_height = max(line_height, max_row_height)

            x_start_row = pdf.get_x()
            y_start_row = pdf.get_y()

            if pdf.get_y() + max_row_height > pdf.h - 20: 
                pdf.add_page()
                pdf.set_font("Arial", "B", 9)
                for i, h in enumerate(headers):
                    pdf.cell(col_widths[i], 8, h, border=1, align='C')
                pdf.ln()
                pdf.set_font("Arial", size=8)
                x_start_row = pdf.get_x()
                y_start_row = pdf.get_y()


            for i, txt in enumerate(row_data):
                pdf.set_xy(x_start_row + sum(col_widths[:i]), y_start_row)
                pdf.multi_cell(col_widths[i], line_height, str(txt), border=0, align='L') 
            
            pdf.set_y(y_start_row)
            pdf.set_x(x_start_row) 
            
            for i in range(len(headers)):
                pdf.cell(col_widths[i], max_row_height, '', border=1, ln=0) 
            pdf.ln() 

            pdf.set_y(y_start_row + max_row_height)


    # Guardar PDF
    pdf.output(ruta_archivo)
    
